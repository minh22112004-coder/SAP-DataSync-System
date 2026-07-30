from __future__ import annotations

import hashlib
import logging
import os
import re
import shutil
import struct
import traceback
import uuid
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta, tzinfo
from decimal import Decimal
from pathlib import Path
from typing import Iterable, Sequence

import pyodbc
from openpyxl import load_workbook

from .config import Settings

LOGGER = logging.getLogger(__name__)

EXPECTED_SOURCE_COLUMN_COUNT = 149
FIRST_BUSINESS_KEY_COLUMN = "Shipping Instructions ID"
SECOND_BUSINESS_KEY_COLUMN = "Unique Number"


@dataclass(frozen=True)
class ImportResult:
    status: str
    import_log_id: str | None = None
    total_rows: int = 0
    inserted_rows: int = 0
    updated_rows: int = 0
    unchanged_rows: int = 0
    deleted_rows: int = 0


class EtlWorker:
    def __init__(self, settings: Settings) -> None:
        self._settings = settings

    def run_cycle(self) -> int:
        source_directories = [Path(self._settings.source_path)]
        if self._settings.upload_path:
            upload_directory = Path(self._settings.upload_path)
            if upload_directory not in source_directories:
                source_directories.append(upload_directory)

        if not source_directories[0].is_dir():
            LOGGER.error("Source directory is not available: %s", source_directories[0])
            return 1

        archive_directory = Path(self._settings.archive_path)
        try:
            archive_directory.mkdir(parents=True, exist_ok=True)
        except OSError:
            LOGGER.exception(
                "Archive directory is not writable: %s", archive_directory
            )
            return 1

        failures = 0
        file_paths: list[Path] = []
        for source_directory in source_directories:
            if not source_directory.is_dir():
                LOGGER.warning("Optional import directory is not available: %s", source_directory)
                continue
            file_paths.extend(source_directory.glob(self._settings.file_pattern))

        file_paths.sort(key=lambda path: (path.name.lower(), str(path.parent).lower()))
        for file_path in file_paths:
            if not file_path.is_file():
                continue
            age_seconds = datetime.now().timestamp() - file_path.stat().st_mtime
            if age_seconds < self._settings.minimum_file_age_seconds:
                LOGGER.info("Waiting for file to stabilize: %s", file_path.name)
                continue

            try:
                file_hash = compute_file_hash(file_path)
                snapshot_path, snapshot_created = create_or_reuse_snapshot(
                    file_path,
                    archive_directory,
                    file_hash,
                    self._settings.timezone,
                )
                LOGGER.info(
                    "%s snapshot: source=%s; archive=%s; sha256=%s",
                    "Created" if snapshot_created else "Reused",
                    file_path.name,
                    snapshot_path.name,
                    file_hash,
                )
                result = self._import_file(snapshot_path, file_hash)
                LOGGER.info(
                    "%s: source=%s; snapshot=%s; total=%s; inserted=%s; updated=%s; "
                    "unchanged=%s; deleted=%s; importLogId=%s",
                    result.status,
                    file_path.name,
                    snapshot_path.name,
                    result.total_rows,
                    result.inserted_rows,
                    result.updated_rows,
                    result.unchanged_rows,
                    result.deleted_rows,
                    result.import_log_id or "-",
                )
            except Exception:
                failures += 1
                LOGGER.exception("Import failed for %s", file_path.name)

        return failures

    def _import_file(self, file_path: Path, file_hash: str) -> ImportResult:
        with pyodbc.connect(self._settings.connection_string, autocommit=False) as connection:
            source_columns = _load_source_columns(connection)
            _validate_database_schema(source_columns)

            if _is_already_completed(connection, file_hash):
                connection.rollback()
                return ImportResult(status="AlreadyCompleted")

            import_log_id = str(uuid.uuid4())
            _create_import_log(
                connection,
                import_log_id,
                file_path.name,
                file_hash,
                self._settings.product,
                self._settings.sales_organization,
                self._settings.soft_delete_enabled,
            )
            connection.commit()

            try:
                _acquire_etl_lock(connection)
                total_rows = self._load_staging(
                    connection,
                    file_path,
                    import_log_id,
                    source_columns,
                )

                if compute_file_hash(file_path) != file_hash:
                    raise OSError(
                        "The Excel file changed while it was being imported. "
                        "The transaction was rolled back."
                    )

                counts = _synchronize(connection, import_log_id)
                if counts.total_rows != total_rows:
                    raise RuntimeError(
                        "Staging and synchronization row counts are inconsistent"
                    )

                _mark_completed(connection, import_log_id, counts)
                connection.commit()
                return ImportResult(
                    status="Completed",
                    import_log_id=import_log_id,
                    total_rows=counts.total_rows,
                    inserted_rows=counts.inserted_rows,
                    updated_rows=counts.updated_rows,
                    unchanged_rows=counts.unchanged_rows,
                    deleted_rows=counts.deleted_rows,
                )
            except Exception as exception:
                connection.rollback()
                try:
                    _mark_failed(connection, import_log_id, exception)
                    connection.commit()
                except Exception:
                    connection.rollback()
                    LOGGER.exception(
                        "Could not mark ImportLog %s as Failed", import_log_id
                    )
                raise

    def _load_staging(
        self,
        connection: pyodbc.Connection,
        file_path: Path,
        import_log_id: str,
        source_columns: Sequence[str],
    ) -> int:
        workbook = load_workbook(
            filename=file_path,
            read_only=True,
            data_only=True,
            keep_links=False,
        )
        try:
            if self._settings.worksheet_name:
                if self._settings.worksheet_name not in workbook.sheetnames:
                    raise ValueError(
                        f"Worksheet '{self._settings.worksheet_name}' was not found"
                    )
                worksheet = workbook[self._settings.worksheet_name]
            else:
                if not workbook.worksheets:
                    raise ValueError("The workbook contains no worksheets")
                worksheet = workbook.worksheets[0]

            rows = worksheet.iter_rows(values_only=True)
            try:
                header = next(rows)
            except StopIteration as exception:
                raise ValueError(f"Worksheet '{worksheet.title}' is empty") from exception

            _validate_header(header, source_columns, worksheet.title)
            first_key_index = source_columns.index(FIRST_BUSINESS_KEY_COLUMN)
            second_key_index = source_columns.index(SECOND_BUSINESS_KEY_COLUMN)
            seen_business_keys: set[bytes] = set()
            batch: list[tuple[object, ...]] = []
            total_rows = 0

            quoted_source_columns = ", ".join(
                _quote_identifier(column) for column in source_columns
            )
            target_columns = (
                "[ImportLogId], [SourceRowNumber], [BusinessKeyHash], [RowHash], "
                + quoted_source_columns
            )
            placeholders = ", ".join("?" for _ in range(4 + len(source_columns)))
            insert_sql = (
                f"INSERT INTO dbo.SapDataStaging ({target_columns}) "
                f"VALUES ({placeholders})"
            )

            cursor = connection.cursor()
            cursor.fast_executemany = True

            for excel_row_number, raw_row in enumerate(rows, start=2):
                values = [
                    _convert_cell_value(raw_row[index] if index < len(raw_row) else None)
                    for index in range(len(source_columns))
                ]
                if not any(value is not None for value in values):
                    continue

                first_key = values[first_key_index]
                second_key = values[second_key_index]
                if not first_key and not second_key:
                    raise ValueError(
                        f"Excel row {excel_row_number} has no business key; at least "
                        f"'{FIRST_BUSINESS_KEY_COLUMN}' or "
                        f"'{SECOND_BUSINESS_KEY_COLUMN}' is required"
                    )

                business_key_hash = _compute_value_hash((first_key, second_key))
                if business_key_hash in seen_business_keys:
                    raise ValueError(
                        f"Excel row {excel_row_number} has a duplicate business key "
                        f"({FIRST_BUSINESS_KEY_COLUMN} + "
                        f"{SECOND_BUSINESS_KEY_COLUMN})"
                    )
                seen_business_keys.add(business_key_hash)

                batch.append(
                    (
                        import_log_id,
                        excel_row_number,
                        business_key_hash,
                        _compute_value_hash(values),
                        *values,
                    )
                )
                total_rows += 1

                if len(batch) >= self._settings.batch_size:
                    cursor.executemany(insert_sql, batch)
                    batch.clear()

            if batch:
                cursor.executemany(insert_sql, batch)

            if total_rows == 0:
                raise ValueError(
                    f"Worksheet '{worksheet.title}' contains no data rows"
                )

            return total_rows
        finally:
            workbook.close()


@dataclass(frozen=True)
class SyncCounts:
    total_rows: int
    inserted_rows: int
    updated_rows: int
    unchanged_rows: int
    deleted_rows: int


def compute_file_hash(file_path: Path) -> str:
    digest = hashlib.sha256()
    with file_path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(64 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest().upper()


def create_or_reuse_snapshot(
    source_path: Path,
    archive_directory: Path,
    file_hash: str,
    timezone: tzinfo,
) -> tuple[Path, bool]:
    """Create an immutable archive copy before importing the source file."""
    archive_directory.mkdir(parents=True, exist_ok=True)

    for existing_path in archive_directory.glob(f"*_{file_hash}.xlsx"):
        if existing_path.is_file() and compute_file_hash(existing_path) == file_hash:
            return existing_path, False

    source_stat = source_path.stat()
    source_stem = _safe_file_stem(source_path.stem)
    if re.search(r"_\d{8}_\d{6}$", source_stem):
        snapshot_stem = source_stem
    else:
        source_modified_at = datetime.fromtimestamp(
            source_stat.st_mtime,
            tz=timezone,
        )
        snapshot_stem = f"{source_stem}_{source_modified_at:%Y%m%d_%H%M%S}"

    snapshot_path = archive_directory / f"{snapshot_stem}_{file_hash}.xlsx"
    temporary_path = archive_directory / f".{snapshot_path.name}.{uuid.uuid4().hex}.tmp"

    try:
        shutil.copyfile(source_path, temporary_path)
        copied_hash = compute_file_hash(temporary_path)
        if copied_hash != file_hash:
            raise OSError(
                "The Excel source changed while its archive snapshot was being created"
            )
        os.replace(temporary_path, snapshot_path)
    finally:
        temporary_path.unlink(missing_ok=True)

    return snapshot_path, True


def _safe_file_stem(value: str) -> str:
    safe_value = re.sub(r"[^A-Za-z0-9._-]+", "_", value).strip("._-")
    return safe_value or "export"


def _load_source_columns(connection: pyodbc.Connection) -> list[str]:
    cursor = connection.cursor()
    rows = cursor.execute(
        """
        SELECT [name]
        FROM sys.columns
        WHERE object_id = OBJECT_ID(N'dbo.SapDataStaging')
          AND [name] NOT IN
              (N'StagingId', N'ImportLogId', N'SourceRowNumber',
               N'BusinessKeyHash', N'RowHash', N'LoadedAt')
        ORDER BY column_id;
        """
    ).fetchall()
    return [str(row[0]) for row in rows]


def _validate_database_schema(source_columns: Sequence[str]) -> None:
    if len(source_columns) != EXPECTED_SOURCE_COLUMN_COUNT:
        raise RuntimeError(
            f"SapDataStaging has {len(source_columns)} source columns; "
            f"expected {EXPECTED_SOURCE_COLUMN_COUNT}"
        )
    if (
        FIRST_BUSINESS_KEY_COLUMN not in source_columns
        or SECOND_BUSINESS_KEY_COLUMN not in source_columns
    ):
        raise RuntimeError("The database does not contain the business key columns")


def _validate_header(
    header: Sequence[object], source_columns: Sequence[str], worksheet_name: str
) -> None:
    if len(header) != len(source_columns):
        raise ValueError(
            f"Worksheet '{worksheet_name}' has {len(header)} columns; "
            f"expected exactly {len(source_columns)}"
        )
    for index, expected in enumerate(source_columns):
        actual = "" if header[index] is None else str(header[index])
        if actual != expected:
            raise ValueError(
                f"Invalid header at Excel column {index + 1}: "
                f"found '{actual}', expected '{expected}'"
            )


def _is_already_completed(connection: pyodbc.Connection, file_hash: str) -> bool:
    row = connection.cursor().execute(
        """
        SELECT CASE WHEN EXISTS
        (
            SELECT 1
            FROM dbo.ImportLog
            WHERE FileHash = ? AND Status = N'Completed'
        ) THEN 1 ELSE 0 END;
        """,
        file_hash,
    ).fetchone()
    return bool(row and row[0])


def _create_import_log(
    connection: pyodbc.Connection,
    import_log_id: str,
    file_name: str,
    file_hash: str,
    product: str,
    sales_organization: str,
    soft_delete_enabled: bool,
) -> None:
    if len(file_name) > 260:
        raise ValueError("The Excel file name exceeds 260 characters")
    connection.cursor().execute(
        """
        INSERT INTO dbo.ImportLog
            (Id, FileName, FileHash, Status, Product, SalesOrganization, SoftDeleteEnabled)
        VALUES (?, ?, ?, N'Processing', ?, ?, ?);
        """,
        import_log_id,
        file_name,
        file_hash,
        product,
        sales_organization,
        soft_delete_enabled,
    )


def _acquire_etl_lock(connection: pyodbc.Connection) -> None:
    row = connection.cursor().execute(
        """
        DECLARE @Result INT;
        EXEC @Result = sys.sp_getapplock
            @Resource = N'SapDataSync:PythonEtlWorker',
            @LockMode = N'Exclusive',
            @LockOwner = N'Transaction',
            @LockTimeout = 0;
        SELECT @Result;
        """
    ).fetchone()
    if row is None or int(row[0]) < 0:
        raise RuntimeError("Another ETL import is already running")


def _synchronize(connection: pyodbc.Connection, import_log_id: str) -> SyncCounts:
    cursor = connection.cursor()
    row = cursor.execute(
        "EXEC dbo.SyncSapData @ImportLogId = ?;", import_log_id
    ).fetchone()
    if row is None:
        raise RuntimeError("dbo.SyncSapData returned no synchronization counts")
    counts = SyncCounts(*(int(value) for value in row[:5]))
    if counts.unchanged_rows < 0:
        raise RuntimeError("Synchronization counts are inconsistent")
    return counts


def _mark_completed(
    connection: pyodbc.Connection, import_log_id: str, counts: SyncCounts
) -> None:
    connection.cursor().execute(
        """
        UPDATE dbo.ImportLog
        SET Status = N'Completed',
            CompletedAt = SYSUTCDATETIME(),
            TotalRows = ?,
            InsertedRows = ?,
            UpdatedRows = ?,
            UnchangedRows = ?,
            DeletedRows = ?,
            ErrorRows = 0,
            ErrorMessage = NULL
        WHERE Id = ?;
        """,
        counts.total_rows,
        counts.inserted_rows,
        counts.updated_rows,
        counts.unchanged_rows,
        counts.deleted_rows,
        import_log_id,
    )


def _mark_failed(
    connection: pyodbc.Connection, import_log_id: str, exception: Exception
) -> None:
    error_message = "".join(
        traceback.format_exception(type(exception), exception, exception.__traceback__)
    )
    connection.cursor().execute(
        """
        UPDATE dbo.ImportLog
        SET Status = N'Failed',
            CompletedAt = SYSUTCDATETIME(),
            ErrorRows = 1,
            ErrorMessage = ?
        WHERE Id = ?;
        """,
        error_message,
        import_log_id,
    )


def _convert_cell_value(value: object) -> str | None:
    if value is None:
        return None
    if isinstance(value, str):
        return value
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S.") + f"{value.microsecond:06d}0"
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d 00:00:00.0000000")
    if isinstance(value, time):
        return value.strftime("%H:%M:%S.") + f"{value.microsecond:06d}"
    if isinstance(value, timedelta):
        return _format_timedelta(value)
    if isinstance(value, bool):
        return "True" if value else "False"
    if isinstance(value, float):
        return str(value)
    if isinstance(value, Decimal):
        return format(value, "f")
    return str(value)


def _format_timedelta(value: timedelta) -> str:
    total_microseconds = int(value.total_seconds() * 1_000_000)
    sign = "-" if total_microseconds < 0 else ""
    total_microseconds = abs(total_microseconds)
    total_seconds, microseconds = divmod(total_microseconds, 1_000_000)
    days, day_seconds = divmod(total_seconds, 86_400)
    hours, remainder = divmod(day_seconds, 3_600)
    minutes, seconds = divmod(remainder, 60)
    day_prefix = f"{days}." if days else ""
    fraction = f".{microseconds:06d}" if microseconds else ""
    return f"{sign}{day_prefix}{hours:02d}:{minutes:02d}:{seconds:02d}{fraction}"


def _compute_value_hash(values: Iterable[str | None]) -> bytes:
    digest = hashlib.sha256()
    for value in values:
        if value is None:
            digest.update(struct.pack("<i", -1))
            continue
        encoded = value.encode("utf-8")
        digest.update(struct.pack("<i", len(encoded)))
        digest.update(encoded)
    return digest.digest()


def _quote_identifier(identifier: str) -> str:
    return "[" + identifier.replace("]", "]]") + "]"
